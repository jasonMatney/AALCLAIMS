# Python 3.6

#---------------------------------------------------------------------------#
#-- Import Modules/Libraries -----------------------------------------------#
#---------------------------------------------------------------------------#
import os, sys, time, glob, h5py, IPython, gdal
import osr
import pandas as pd
import numpy as np
import geopandas as gpd
import pathlib as pl
from pandas.api.types import is_numeric_dtype
from osgeo import gdal
import rasterio
from io import BytesIO
from openpyxl import load_workbook
from matplotlib import pyplot as plt
from matplotlib.ticker import FuncFormatter
from IPython.display import display, Markdown, Latex
import json

pd.set_option('display.max_columns', 20)
plt.rcParams.update({'font.size': 14})

#---------------------------------------------------------------------------#
#-- Type Hints for Functions -----------------------------------------------#
#---------------------------------------------------------------------------#
printextra = "Does work but only reports results through prints"
pd_df, pandas_DF = pd.DataFrame, pd.DataFrame
mpl_figure = plt.subplots
gdal_data = 'rasterband, geotransform, and gdal raster objects'

#---------------------------------------------------------------------------#
#-- Basic Formatting -------------------------------------------------------#
#---------------------------------------------------------------------------#

def IfDirExistsElseCreate(path_dir: str, print_stat: bool = False) -> print:
    """Trys to create a folder directory if it does not exist."""
    
    try: os.makedirs(path_dir)
        
    except OSError: 
        if print_stat: display(Markdown(f"The directory '{path_dir}' " \
                                        + "already exists"))
    else: 
        if print_stat: display(Markdown(f"Created '{path_dir}'"))
            
    return None

#---------------------------------------------------------------------------#
#-- Working with GDAL ------------------------------------------------------#
#---------------------------------------------------------------------------#

def query(x: float, y: float, gt: any, rb: any) -> float:
    """Queries one specific cell in the rasterband given an x, y in the 
    geotransform
    """
    px = int((x-gt[0]) / gt[1])   
    py = int((y-gt[3]) / gt[5])   
    return rb.ReadAsArray(px,py,1,1)[0][0]

#---------------------------------------------------------------------------#
#-- Read from AWS ----------------------------------------------------------#
#---------------------------------------------------------------------------#

try:
    import boto3
    s3 = boto3.resource('s3')
    def s3List(bucketName, prefixName, nameSelector, fileformat):
        """
            This function takes an S3 bucket name and prefix (flat directory path) and returns a list of GeoTiffs.
                This function utilizes boto3's continuation token to iterate over an unlimited number of records.

            BUCKETNAME -- A bucket on S3 containing GeoTiffs of interest
            PREFIXNAME -- A S3 prefix.
            NAMESELECTOR -- A string used for selecting specific files. E.g. 'SC' for SC_R_001.tif.
            FILEFORMAT -- A string variant of a file format.
        """
        # Set the Boto3 client
        s3_client = boto3.client('s3')
        # Get a list of objects (keys) within a specific bucket and prefix on S3
        keys = s3_client.list_objects_v2(Bucket=bucketName, Prefix=prefixName)
        # Store keys in a list
        keysList = [keys]
        # While the boto3 returned objects contains a value of true for 'IsTruncated'
        while keys['IsTruncated'] is True:
            # Append to the list of keys
            # Note that this is a repeat of the above line with a contuation token
            keys = s3_client.list_objects_v2(Bucket=bucketName, Prefix=prefixName,
                                             ContinuationToken=keys['NextContinuationToken'])
            keysList.append(keys)

        # Create a list of GeoTiffs from the supplied keys
        #     While tif is hardcoded now, this could be easily changed.
        pathsList = []
        for key in keysList:
            paths = ['s3://'+ bucketName + '/' + elem['Key'] for elem in key['Contents'] \
                         if elem['Key'].find('{}'.format(nameSelector))>=0 and elem['Key'].endswith(fileformat)]
            pathsList = pathsList + paths

        return pathsList

    def getTifData(s3path: s3.Object) -> gdal_data:
        """Read a raster from S3 into memory and get attributes"""
        
        image_data = BytesIO(s3path.get()['Body'].read())
        tif_inmem = "/vsimem/data.tif" #Virtual Folder to Store Data
        gdal.FileFromMemBuffer(tif_inmem, image_data.read())
        print(s3path)
        src = gdal.Open(tif_inmem)  
        rb, gt = src.GetRasterBand(1), src.GetGeoTransform()
        
        return rb, gt, src

    def s3Attributes(s3path: str, replace_str: str = '', rtype: str = '') -> any:
        """Creates the S3 object to write or read from AWS bucket"""
        
        parent = s3path.split(':')[0]
        name = s3path.split(':')[1].split('//')[-1].replace(replace_str,'')
        s3path = s3.Object(s3path.split(':')[0], s3path.split(':')[1])
        
        if rtype == 'NAME': 
            return name
        elif rtype == 'PARENT': 
            return parent
        elif rtype == 'S3PATH': 
            return s3path
        else: 
            return parent, name, s3path

    def get_files(bucketname: str, prefixname: str, textstring: str) -> list:
        """For navigating "folder" structures in bucket, Get list of files from 
        bucket/prefix in S3 https://stackoverflow.com/questions/35803027/
        retrieving-subfolders-names-in-s3-bucket-from-boto3
        """
        wselist = []
        s3 = boto3.resource('s3')
        bucket = s3.Bucket(name=bucketname)
        FilesNotFound = True
        
        for obj in bucket.objects.filter(Prefix=prefixname):
            if textstring in str(obj) and 'xml' not in str(obj):
                wselist.append('{0}:{1}'.format(bucket.name, obj.key))
            FilesNotFound = False
            
        if FilesNotFound:
            print("ALERT", "No file in {0}/{1}".format(bucket, prefixname))
        return wselist 

except:
    print('Verify Boto3/AWS installation, credentials, & access')


##--Pick up WSE from local/remote
def GetWSE_JobID(tif, use_parts=list(), jobIDString = '',topo=False):
    if topo:
        return 'Groundelev'
    else:
        path_parts = tif.split('/') # Use for s3/unix

        if len(path_parts)==1:
            path_parts = pl.Path(tif).parts

        jobID=jobIDString
        for part in use_parts:
            jobID+='{}-'.format(path_parts[part])
            
        return jobID[:-1]

def CheckProjs(ds, gdf0, bruteforceProj=None):
    try:
        tif_crs = osr.SpatialReference(ds.GetProjectionRef()).ExportToProj4()
        tif_crs = rasterio.crs.CRS.from_proj4(tif_crs).to_wkt()
        if gdf0.crs==tif_crs:
            return gdf0
        
        elif bruteforceProj:
            gdf = gdf0.to_crs(bruteforceProj)
            return gdf

        elif tif_crs!=gdf0.crs:
            gdf = gdf0.to_crs(tif_crs)
            return gdf
    except:
        return None
    

##--Pick up WSE from local/remote
def GetWSE(tif, gdf, json_dir, bldgidfield, null_value, tifs_to_json=list(), use_parts=list(),
    jobIDString = '',istopo=False, readlocal=False,writelocal=True, mod_rate=2, nameSelector=None, bucketName=None):
    """This function (and nested functions) needs cleaning, documenting, and general improvements"""
    
    jobID = GetWSE_JobID(tif, use_parts, jobIDString, istopo)
    
    
    if readlocal and writelocal:
        json_file = pl.Path(json_dir)/'{}.json'.format(jobID)
        
    elif (not readlocal) and (writelocal):
        tifkey = tif.replace('s3://{}/'.format(bucketName),'')
        s3tif = s3.Object(bucketName, tifkey)
        json_file = pl.Path(json_dir)/'{}.json'.format(jobID)
    
    else: # all on aws
        tifkey = tif.replace('s3://{}/'.format(bucketName),'')
        s3tif = s3.Object(bucketName, tifkey)
        json_file =  "s3://{}/".format(bucketName) +json_dir+'/'+jobID+'.json'
        
    if str(json_file) not in tifs_to_json or istopo:
        if readlocal:
            rb, gt, ds = getTifData_local(tif)
            #gdf = CheckProjs(ds, gdf)
            
        else:
            rb, gt, ds = getTifData(s3tif)
            #gdf = CheckProjs(ds, gdf)
            
        results = readStructureData(gdf, ds, rb, gt, bldgidfield, null_value)
        data = json.dumps(results)
        return data, json_file
        
    else: 
        return None, None
        


def WriteWSE_Json(json_file, data, writelocal=False):
    if data==None:
        return None
    
    if writelocal:    
        with open(json_file, 'w') as f:  
            json.dump(data, f)

    else:
        client = boto3.client('s3')
        response = client.put_object(Body=json.dumps(results), Bucket=bucketName, Key='/'.join([json_dir,jobID+'.json']))
    return None


##------------------------
def readStructureData(
        gdf: gpd.GeoDataFrame, ras: gdal_data, rb: gdal_data, 
        gt: gdal_data, bldgidfield: str, null_value: float) -> list:
    """Read Structures and return a list of IDs and WSE values"""
    results = {}
            
    for i, idx in enumerate(gdf.index):
        uniqueId = gdf.loc[idx,bldgidfield]
        bldg = gdf.loc[idx,'geometry']
        x, y = bldg.x, bldg.y
        pixel_value = query(x,y,gt,rb)
        if str(pixel_value)==str(null_value):
            results[uniqueId] = ""
        else:
            results[uniqueId] = str(pixel_value)
    
    return results

def determineRunList(ftype: str, specific_area: str, 
                     num_of_breaches: int, breach_start_num: int) -> list:
    """Returns the list of folders that will be
    searched in the flood scenario.
    """
    runs = []
    
    if num_of_breaches:
        runs = [f'Base{specific_area}']
        for i in range(num_of_breaches):
            runs.append(f'Breach{i + breach_start_num}')
    else:
        runs = [ftype]
        
    return runs

#---------------------------------------------------------------------------#
#--Read Local WSE ----------------------------------------------------------#
#---------------------------------------------------------------------------#
def getTifData_local(tif_path: str) -> gdal_data:
    """Read a local raster and return the GDAL objects"""
    src = gdal.Open(tif_path)
    rb, gt = src.GetRasterBand(1), src.GetGeoTransform()
    return rb, gt, src

#---------------------------------------------------------------------------#
#-- QC Data and Control File -----------------------------------------------#
#---------------------------------------------------------------------------#
def GetHazusData(ControlFilePath: str, HzJSON: str) -> printextra:
    """Writes Hazus Building Depth Damage Functions to the control
    file if it is not already there.
    """
    hz_DepthDmgFnId = pd.read_json(HzJSON, orient = 'index')
    HzSheetName = "HazusDepthDmgFns"
    wb = load_workbook(ControlFilePath,read_only=True)
    if HzSheetName in wb.sheetnames:
        print(f'{HzSheetName} already exists')
    else:
        with pd.ExcelWriter(ControlFilePath,engine='openpyxl') as writer:
            writer.book = load_workbook(ControlFilePath)
            hz_DepthDmgFnId.to_excel(writer,"HazusDepthDmgFns")
        print(f'{HzSheetName} created')
    return None

def CheckColumnsFromCF(cf_pathlocs: pd_df, cf_shpdetails: pd_df) -> dict:
    """Checks the Shapefile for Columns Listed in 
    Control File
    """
    SPL = 'Shapefile Path Location'
    
    for i in cf_pathlocs.index:
        shppath = pl.PurePath(cf_pathlocs.loc[i,SPL])
        NoIssuesFound = True
        print(f'Checking {shppath.name}')
        gdf = gpd.read_file(str(shppath))
        
        for j in cf_shpdetails.index:
            ctype = cf_shpdetails.loc[j,'Shapefile Column ID']
            cname = cf_shpdetails.loc[j,'Shapefile Column Names']
            
            if cname not in gdf.columns.tolist():
                if cf_shpdetails.loc[j,'Required'] == 'Yes': 
                    NoIssuesFound = False
                print(f' - Required: {cf_shpdetails.loc[j,"Required"]}'
                      + f'\n   Not Found: Column Type "{ctype}" with' \
                      + f' name "{cname}" ')
            else:
                if 'column_dict' not in locals():
                    column_dict = {ctype:cname}
                else:
                    column_dict.update({ctype:cname})
                    
        print(f'  {shppath}')
        if NoIssuesFound: 
            print(f' - No issues found for {shppath.name}')
            print(f' - {column_dict}')
            
    return column_dict

def CheckWSEColsMatchWeights(cf_weights: pd_df, 
                             cf_pathlocs: pd_df) -> printextra:
    """Checks the Columns in the WSE csv and compares them to the
    wieght runs
    """
    weightruns = cf_weights.index.tolist()
    WFPL = 'WSE File Path Locations'
    
    for i in cf_pathlocs.index:
        wsepath = pl.PurePath(cf_pathlocs.loc[i,WFPL])
        df_wse = pd.read_csv(str(wsepath),index_col=0)
        notfound = []
        
        for w in weightruns:
            if w not in df_wse.columns.tolist():
                notfound.append(w)
                
        if notfound == []:
            notfound = 'No Issues Found'
            print(f'{wsepath.name}\n  {notfound}')
        else:
            print(f'{wsepath.name} missing:\n  {notfound}')
    
    return None

def CheckDamageCategories(cf_dc: pd_df) -> dict:
    """Checks the Damage Categories and returns them in a ditionary"""
    HzDFnId = 'Hazus Damage Function IDs'
    cf_damagecats = cf_dc.copy()
    
    for i in cf_damagecats.index: 
        dmgs = cf_damagecats.loc[i,HzDFnId].split(',')
        dmgs_int = [int(DmgfnId) for DmgfnId in dmgs]
        cf_damagecats.loc[i,HzDFnId] = dmgs_int

    for cat in cf_damagecats.index:
        if 'catgroups' not in globals():
            catgroups = {cat:cf_damagecats.loc[cat,HzDFnId]}
        else:
            catgroups.update({cat:cf_damagecats.loc[cat,HzDFnId]})
    
    return catgroups

#---------------------------------------------------------------------------#
#-- Prep Data for Calculations ---------------------------------------------#
#---------------------------------------------------------------------------#
def hazusID_to_depth(df: pandas_DF) -> pandas_DF:
    """Formats Hazus Depth In Structure vs Damages table to be readable 
    for this script.
    """
    rawCol, rawDIS, newDIS = df.columns.tolist(), [], [] 
    for col in rawCol:
        if col[0] == 'm' or col[0] == 'p':
            rawDIS.append(col)
            newcol = int(col.replace('m','-').replace('p',''))
            newDIS.append(newcol)
    for i, col in enumerate(rawDIS):
        df.rename(columns={col:newDIS[i]}, inplace=True)
    return df

def aggregate_ddf_curves(df: pandas_DF, curve_groups: dict, 
                         custom_depths: list = False, 
                         plot: bool = True) -> pandas_DF:
    '''curve_groups is a dictionary categrizing a list of damage 
    functions to aggregate e.g. "Category1": [1,2,3].
    '''
    depths_in_curves = custom_depths if custom_depths else list(range(-4,25))
    df_agg = pd.DataFrame()
    for group in curve_groups.keys():
        dfc = df.loc[curve_groups[group]][depths_in_curves].T
        occ_type =  df['Occupancy'].loc[curve_groups[group]].unique()[0]
        df_agg[group] = dfc.mean(axis=1)
        if plot:
            fig, ax = plt.subplots(figsize=(22,4))
            for idx in dfc.columns:
                ax.plot(dfc[idx], linestyle='--', label =str(idx))
            ax.plot(dfc.mean(axis=1), label='Mean', color='black')
            ax.set_title(f'Raw Depth Damage curves for {group}' \
                         + f'\n({occ_type})',fontsize=20)
            ax.legend()
            ax.grid()
            ax.set_xlabel('Depth (ft)', fontsize=16)
    return df_agg

#---------------------------------------------------------------------------#
#-- Losses Calculated ------------------------------------------------------#
#-----------------------------------------------------------------------------#
def CalcLosses(df_parcel_wse: pandas_DF, df_agg: pandas_DF, 
               structure_cols: dict, Return_DFs: str = 'ALL') -> pandas_DF:
    """For every building it determines depth in structure and the 
    correlated damages for each event. 
    """
    start_time = time.time()
    
    # Create depths, Damage Percent, and Damage Losses dataframes
    df_parcel_depth = pd.DataFrame(index = df_parcel_wse.index)
    for col in df_parcel_wse.columns.tolist():
        if col not in list(structure_cols.keys()):
            if is_numeric_dtype(df_parcel_wse[col]):
                df_parcel_depth[col] = df_parcel_wse[col] \
                                       - df_parcel_wse[structure_cols['Ground Elevation']] \
                                       - df_parcel_wse[structure_cols['First Floor Height']]
    df_DP = df_parcel_depth.copy()
    df_DL = df_parcel_depth.copy()
    df_DLinsurance = df_parcel_depth.copy()
    
    # Create Group index to analyze different Damage Codes (if used)
    ParcelBldgGroups = []
    for dckey in df_agg.columns.tolist():
        cat_key = df_parcel_wse[df_parcel_wse[structure_cols['Damage Code']] == dckey].index.tolist()
        ParcelBldgGroups.append(cat_key)
    
    # Perform the Damage Calculations
    for col in df_DP.columns.tolist():
        for i, cat_key in enumerate(ParcelBldgGroups):
            dckey = df_agg.columns.tolist()[i]
            depthindex = df_agg.index.tolist()
            lossindex = df_agg[dckey].values.tolist()
            interpolateLosses = lambda x: np.interp(x, depthindex, lossindex, 
                                                    left = 0, right = lossindex[-1])
            df_DP.loc[cat_key, col] = df_parcel_depth.loc[cat_key,col].apply(interpolateLosses)
            df_DL.loc[cat_key, col] = df_parcel_wse.loc[cat_key, structure_cols['Building Limit']] \
                                      * ( df_DP.loc[cat_key, col] / 100.0 )
            df_DLinsurance.loc[cat_key, col] = df_DL.loc[cat_key, col] \
                                               - df_parcel_wse.loc[cat_key, structure_cols['Building Deduction']]
    
    # Filter values for non-affected cells
    df_parcel_depth[df_parcel_depth < -10] = np.nan
    df_DLinsurance[df_DLinsurance < 0] = 0
    
    # Now Check what the user wants to return
    print(f"CalcLosses Total Time: {((time.time()-start_time)/60):0.2f} minutes")
    if Return_DFs=='ALL':
        value = {'loss_pct': df_DP, 'loss_usd': df_DL, 
                 'depths':df_parcel_depth, 'loss_insurance': df_DLinsurance}
        return value
    elif Return_DFs=='INSURANCE_DEPTH_AND_PERCENT':
        value = {'loss_pct': df_DP, 'depths':df_parcel_depth, 
                 'loss_insurance': df_DLinsurance}
        return value
    elif Return_DFs=='PERCENT_ONLY': 
        return {'loss_pct': df_DP}
    elif Return_DFs=='LOSSES_ONLY': 
        return {'loss_usd': df_DL}
    elif Return_DFs=='INSURANCE_ONLY': 
        return {'loss_insurance': df_DLinsurance}

#---------------------------------------------------------------------------#
#-- AAL Caclulated from Losses and Weights ---------------------------------#
#---------------------------------------------------------------------------#
def calcAAL_Prob(df_losses: pandas_DF, df_weights: pandas_DF, 
                 model_weight: str = 'RunWeight',
                 method: str = 'Center') -> pandas_DF:
    """Calculates the Average Annual Loss (AAL) for each building, with
    each event being given a weight.
    AAL_method Inputs:
    - "Average" (Traditional Method): This is the traditional method used
    by FEMA to describe damage calculations.
    - "Center": This method assumes the damages are at the center of the
    fequency being used and assumes it is the average. This is the default
    """
    start_time, AAL, prob_weights = time.time(), 'AAL', 'Prob Weights'
    
    ## Sort weights into Cumulative Sum descending order
    df_weights_sort = df_weights.sort_values(by=[model_weight], ascending=False).copy()
    
    # Create Zero Columns for total AAL calculations
    df_LossesWithAAL = df_losses.fillna(value=0).copy()
    df_LossesWithAAL[AAL] = 0
    
    if method == 'Center':
        df_LossesWithAAL[AAL] = 0
        for col in df_weights_sort.index:
            wval = df_weights_sort.loc[col,model_weight]
            df_LossesWithAAL[AAL] = (df_LossesWithAAL[col] * wval) + df_LossesWithAAL[AAL]
        try:
            mprint('## AAL Methodology Used: "Center"\n\nPlease see the ' \
                   + '[Dewberry probmod-tools wiki](https://github.com/Dewberry/probmod-' \
                   + 'tools/wiki/Average-Annual-Losses-(AAL))' \
                   + 'to see how calculations were performed.')
        except:
            print('AAL Methodology Used: "Center"\nProbabilistic Preferred')
            
    elif method == 'Average':
        for i, col in enumerate(df_weights_sort.index):
            if i == 0:
                prev_wcol = col
                prev_wval = df_weights_sort.loc[col,model_weight]
            elif i < df_weights_sort.shape[0]:
                curr_wcol = col
                curr_wval = df_weights_sort.loc[col,model_weight]
                prev_AAL = df_LossesWithAAL[AAL]
                df_LossesWithAAL[AAL] = ((df_LossesWithAAL[prev_wcol] + df_LossesWithAAL[curr_wcol]) 
                                        * prev_wval) / 2 + prev_AAL
                # Then Set Previous to the current for next calc
                prev_wcol = curr_wcol
                prev_wval = curr_wval

        # One last hurrah for the equation, since the last value is just multiplied by it's Return Period
        curr_wcol = col
        curr_wval = df_weights_sort.loc[col,model_weight]
        prev_AAL = df_LossesWithAAL[AAL]
        df_LossesWithAAL[AAL] = (df_LossesWithAAL[curr_wcol] * curr_wval) + prev_AAL
        try:
            mprint('## AAL Methodology Used: "Average" (Traditional Method)\n\nPlease see the ' \
                   + '[Dewberry probmod-tools wiki](https://github.com/Dewberry/probmod-tools/wiki/Average-Annual-Losses-(AAL))' \
                   + 'to see how calculations were performed.')
        except:
            print('AAL Methodology Used: "Average"\nHazus Default')
    
    # Function Finished. return dict of products for user
    print('CalcAAL Total Time: {:0.2f} minutes'.format((time.time()-start_time)/60))
    
    return {'losses_with_aal':df_LossesWithAAL,'weights_sorted':df_weights_sort}

def DisplayStatsForAAL(df_losses_withaal: pandas_DF) -> print:
    """Displays the Statistics for all the mean and sum of the 
    Average Annuallized Losses
    """
    aal_avg = '${:,.2f}'.format(df_losses_withaal['AAL'].mean())
    aal_sum = '${:,.2f}'.format(df_losses_withaal['AAL'].sum())
    display(Markdown(f'# Average AAL: {aal_avg}'))
    display(Markdown(f'# Total AAL: {aal_sum}'))
    return None

#---------------------------------------------------------------------------#
#-- Data Formatting For Plotting Results -----------------------------------#
#---------------------------------------------------------------------------#
def SortForTableauTbls(df_depths: pandas_DF,df_losses: pandas_DF,
                       df_RP: pandas_DF) -> pandas_DF:
    """Creates a Damage and Loss index for each buidling with individual
    Probabilities of Occurance
    """
    start_time = time.time()
    bldg_dfs = []
    for idx in df_depths.index:
        df0 = pd.DataFrame(df_depths.loc[idx]).rename({idx:'FloodDepth'},axis=1).replace(np.nan,0)
        df1 = pd.DataFrame(df_losses.loc[idx]).rename({idx:'USDLosses'},axis=1).replace(np.nan,0)
        df = pd.concat([df0,df1,df_RP['RunWeight']],axis=1).sort_values(by=['USDLosses'])
        zeroloss_idx = df['USDLosses'][df['USDLosses']==0].index.tolist()
        df['USDLosses'].loc[zeroloss_idx] = np.nan
        df['FloodDepth'].loc[zeroloss_idx] = np.nan
        IndProbWeights = []
        for i, run in enumerate(df.index):
            if i == 0:
                probweight = df['RunWeight'].sum()
                IndProbWeights.append(probweight)
                prevrunwght = df.loc[run,'RunWeight']
            else:
                probweight = probweight - prevrunwght
                IndProbWeights.append(probweight)
                prevrunwght = df.loc[run,'RunWeight']
        df['BldgProbWeight'] = IndProbWeights
        df = df.drop(columns='RunWeight').reset_index(drop=True)
        bldg_dfs.append(df)
    df_sortindviduals = pd.concat(bldg_dfs,axis=1,keys=df_depths.index)
    print('SortForTableauTbls Run Time: {:0.3f} minutes'.format((time.time()-start_time)/60))
    return df_sortindviduals

#---------------------------------------------------------------------------#
#-- DIAGONOSTIC FUNCTIONS --------------------------------------------------#
#---------------------------------------------------------------------------#

#-----------------------------------------------------------------------------#
#-- AAL Summary Revised ------------------------------------------------------#
#-----------------------------------------------------------------------------#

def findAALfiles4summaries(outputs: pl.Path) -> list:
    """Get list of AAL Binaries from the desingated folder"""
    aal_files = glob.glob(str(outputs / 'AAL__*'), recursive = True)
    mprint(f'## Found {len(aal_files)} AAL Binaries')
    return aal_files

def createDFofAALpaths(aal_files: list) -> pd.DataFrame:
    """Creates a readable dataframe with some basic info seperated out from the 
    AAL binaries found in the aal_files input. Returns a Pandas DataFrame.
    """
    FullPaths, Name, Scenario, Structure, AAL = 'FullPaths', 'Name', 'Scenario', 'Structure', 'AAL'
    df_aalpaths = pd.DataFrame(aal_files,columns=[FullPaths])
    
    getplname = lambda x: pl.Path(x).name.replace('.pkl','').replace('AAL__','')
    getfraggroup = lambda x: f'{pl.Path(x).name.split("__")[1]}'
    getstructures = lambda x: pl.Path(x).name.split('__')[-1].replace('.pkl','')
    
    df_aalpaths[Name] = df_aalpaths[FullPaths].apply(getplname)
    df_aalpaths[Scenario] = df_aalpaths[FullPaths].apply(getfraggroup)
    df_aalpaths[Structure] = df_aalpaths[FullPaths].apply(getstructures)
    df_aalpaths.set_index(Name,drop=True,inplace=True)
    
    print(df_aalpaths.shape)
    return df_aalpaths

def createDFofAALbinaries(df_aalpaths: pd.DataFrame) -> pd.DataFrame:
    """"""
    FullPaths, Name, Scenario, Structure, AAL = 'FullPaths', 'Name', 'Scenario', 'Structure', 'AAL'
    groupdflist = []
    for aalname in df_aalpaths.index:
        aalpath = df_aalpaths.loc[aalname,FullPaths]
        scenname = df_aalpaths.loc[aalname,Scenario]
        df_aal = pd.read_pickle(aalpath)
        df_aal.rename(columns={AAL:scenname}, inplace = True)
        groupdflist.append(df_aal.copy())
    df_aal = pd.concat(groupdflist,axis=1)
    print(df_aal.shape)
    return df_aal

def createAALstats(df_aal: pd.DataFrame) -> pd.DataFrame:
    """Group together some basic statistics from each AAL Group"""
    stat_array = np.array([df_aal.mean(),df_aal.median(),df_aal.min(),df_aal.max()]).T
    stat_cols = ['Average','Median','Minimum','Maximum']
    stat_index = df_aal.mean().index
    df_stats = pd.DataFrame(stat_array,columns=stat_cols,index=stat_index)
    return df_stats

def readControlFileForBoxplots(ControlFilePath: pl.Path, printnames: bool = True) -> pd.DataFrame:
    """"""
    df_controlfile = pd.read_excel(ControlFilePath,sheet_name='Path Locations',index_col=0)
    drange, dscens, dnames_ = [], [], []
    for i, idx in enumerate(df_controlfile.index):
        plotname = df_controlfile.loc[idx,'Proper Name']
        drange.append(i+1)
        dscens.append(idx)
        dnames_.append(plotname)
    if printnames:
        print(drange)
        print(dscens)
        print(dnames_)
    return df_controlfile, drange, dscens, dnames_
